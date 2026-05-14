"""
每日学习计划桌面应用
- 读取 Excel "研0提升计划表"，展示今日及之后未完成任务
- 点击打勾 → 实时同步 Excel
- 紧凑模式：拖到顶部折叠
"""

import sys
import ctypes
from ctypes import wintypes
from datetime import datetime, date
import openpyxl
import threading

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QCheckBox, QScrollArea, QFrame,
    QGraphicsDropShadowEffect, QMessageBox
)
from PySide6.QtCore import Qt, QTimer, QRect, QRectF, QPoint, QMetaObject, Signal
from PySide6.QtGui import (
    QFont, QPainter, QColor, QPen, QPainterPath, QMouseEvent, QLinearGradient, QIcon
)

# ==================== Excel Config ====================

EXCEL_PATH = r"D:\SoftWare\BaiduNetdisk\BaiduSyncdisk\学习计划.xlsm"
SHEET_NAME = "研0提升计划表"
REFRESH_INTERVAL_MS = 60_000  # 60秒自动刷新

# ==================== Windows Acrylic Blur ====================

class ACCENT_POLICY(ctypes.Structure):
    _fields_ = [
        ("AccentState", ctypes.c_uint),
        ("AccentFlags", ctypes.c_uint),
        ("GradientColor", ctypes.c_uint),
        ("AnimationId", ctypes.c_uint),
    ]

class WINCOMPATTRDATA(ctypes.Structure):
    _fields_ = [
        ("Attribute", ctypes.c_int),
        ("Data", ctypes.POINTER(ACCENT_POLICY)),
        ("SizeOfData", ctypes.c_size_t),
    ]

SetWindowCompositionAttribute = ctypes.windll.user32.SetWindowCompositionAttribute
SetWindowCompositionAttribute.argtypes = (wintypes.HWND, ctypes.POINTER(WINCOMPATTRDATA))
SetWindowCompositionAttribute.restype = ctypes.c_bool

def enable_acrylic_blur(hwnd):
    try:
        accent = ACCENT_POLICY()
        accent.AccentState = 4
        accent.AccentFlags = 2
        accent.GradientColor = 0x01000000
        data = WINCOMPATTRDATA()
        data.Attribute = 19
        data.Data = ctypes.pointer(accent)
        data.SizeOfData = ctypes.sizeof(accent)
        return SetWindowCompositionAttribute(hwnd, data)
    except Exception:
        return False

# ==================== Task Type Colors ====================

TYPE_COLORS = {
    "机器学习": "#3B82F6",
    "嵌入式": "#F59E0B",
    "问题": "#EF4444",
    "基础": "#10B981",
    "政治": "#8B5CF6",
    "数学": "#8B5CF6",
    "英语": "#EC4899",
}

def get_type_color(task_type):
    for key, color in TYPE_COLORS.items():
        if key in str(task_type):
            return color
    return "#6366F1"  # default indigo

def get_weekday_name(serial):
    """将Excel星期序列号转换为中文星期"""
    try:
        # Excel serial: 46125 = Monday (1900-01-01 = 1, 但Excel有个bug把1900-02-29算进去了)
        # 所以46125实际是 2026-04-13 = Monday
        d = date(2026, 1, 1) + __import__("datetime").timedelta(days=int(serial) - 1)
        weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        return weekdays[d.weekday()]
    except:
        return str(serial)

# ==================== Excel Data Layer ====================

def load_tasks():
    """加载所有任务，返回字典列表"""
    tasks = []
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, keep_vba=True)
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=2):
            date_val = row[0].value  # A列: 日期
            task_val = row[2].value  # C列: 任务
            if date_val and task_val and isinstance(date_val, datetime):
                task_date = date_val.date()
                is_done = row[4].value is not None
                tasks.append({
                    "date": task_date,
                    "weekday": get_weekday_name(row[1].value),
                    "task": task_val,
                    "type": row[3].value or "",
                    "done": is_done,
                    "done_time": row[5].value,
                    "row_num": row[0].row,
                })
        wb.close()
    except FileNotFoundError:
        print(f"文件未找到: {EXCEL_PATH}")
    except Exception as e:
        print(f"读取Excel失败: {e}")
    return tasks

def mark_done(row_num, done=True):
    """标记任务完成状态，同步到Excel"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)
        ws = wb[SHEET_NAME]
        cell_e = ws.cell(row=row_num, column=5)  # E列: 完成情况
        cell_f = ws.cell(row=row_num, column=6)  # F列: 完成时间
        if done:
            cell_e.value = "√"
            cell_f.value = datetime.now()
        else:
            cell_e.value = None
            cell_f.value = None
        wb.save(EXCEL_PATH)
        wb.close()
        return True
    except PermissionError:
        print(f"[ERROR] 保存失败: Excel文件被其他程序占用")
        QMessageBox.warning(None, "文件被占用", "Excel 文件已被其他程序打开，请关闭后重试。")
        return False
    except Exception as e:
        print(f"[ERROR] 保存失败: {e}")
        return False

# ==================== Modern Button ====================

ACCENT = "#7C3AED"

class ModernButton(QPushButton):
    def __init__(self, text="", color="#FFFFFF", bg_hover="rgba(255,255,255,0.12)",
                 bg_pressed="rgba(255,255,255,0.06)", radius=8, padding="6px 16px", font_size="12px"):
        super().__init__(text)
        self.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {color};
                border: 1px solid rgba(255,255,255,0.15);
                border-radius: {radius}px;
                padding: {padding};
                font-size: {font_size};
            }}
            QPushButton:hover {{
                background: {bg_hover};
                border: 1px solid rgba(255,255,255,0.3);
            }}
            QPushButton:pressed {{
                background: {bg_pressed};
            }}
            QPushButton:disabled {{
                color: rgba(255,255,255,0.2);
                border: 1px solid rgba(255,255,255,0.05);
            }}
        """)
        self.setCursor(Qt.PointingHandCursor)

# ==================== Date Section Widget ====================

class DateSection(QFrame):
    """单个日期区块，含标题和任务列表"""
    def __init__(self, date_str, tasks, on_toggle, is_today=False, parent=None):
        super().__init__(parent)
        self.tasks = tasks
        self.on_toggle = on_toggle
        self.collapsed = not is_today  # 非今日默认折叠
        self._setup_ui(date_str, is_today)

    def _check_changed(self, done):
        """TaskItem勾选变化时更新统计显示"""
        total = len(self.tasks)
        done_count = sum(1 for t in self.tasks if t["done"])
        undone = total - done_count
        if undone == 0:
            self.stat_label.setText(f"({done_count}/{total}) 全部完成")
            self.stat_label.setStyleSheet("color: rgba(16,185,129,0.8); font-size: 11px;")
        else:
            self.stat_label.setText(f"({done_count}/{total}) {undone}个未完成")
            self.stat_label.setStyleSheet("color: rgba(251,191,36,0.8); font-size: 11px;")

    def _setup_ui(self, date_str, is_today=False):
        total = len(self.tasks)
        done = sum(1 for t in self.tasks if t["done"])
        undone = total - done

        if undone == 0:
            stat_text = f"({done}/{total}) 全部完成"
            stat_color = "rgba(16,185,129,0.8)"
        else:
            stat_text = f"({done}/{total}) {undone}个未完成"
            stat_color = "rgba(251,191,36,0.8)"

        self.setStyleSheet("background: transparent;")

        # 顶层垂直布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(12, 8, 12, 0)
        main_layout.setSpacing(0)

        # 日期标题行（可点击折叠）
        header = QFrame()
        header.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 rgba(80,60,140,80), stop:1 rgba(60,50,100,60));
            border-radius: 8px;
            border: 1px solid rgba(100,80,180,60);
        """)
        header_h = QHBoxLayout(header)
        header_h.setContentsMargins(12, 8, 12, 8)
        header_h.setSpacing(10)

        self.toggle_btn = QLabel("▼" if not self.collapsed else "▶")
        self.toggle_btn.setStyleSheet("color: rgba(255,255,255,0.4); font-size: 10px;")

        date_label = QLabel(date_str)
        date_label.setStyleSheet("color: rgba(255,255,255,0.8); font-size: 12px; font-weight: 500;")

        self.stat_label = QLabel(stat_text)
        self.stat_label.setStyleSheet(f"color: {stat_color}; font-size: 11px;")

        header_h.addWidget(self.toggle_btn, 0, Qt.AlignCenter)
        header_h.addWidget(date_label, 0, Qt.AlignLeft | Qt.AlignVCenter)
        header_h.addWidget(self.stat_label, 0, Qt.AlignRight | Qt.AlignVCenter)
        header_h.addStretch()

        # 点击整个 header 切换折叠
        header.mousePressEvent = lambda e: self._toggle()
        self.toggle_btn.mousePressEvent = lambda e: self._toggle()
        date_label.mousePressEvent = lambda e: self._toggle()
        self.stat_label.mousePressEvent = lambda e: self._toggle()

        # 任务列表容器
        self.task_container = QFrame()
        self.task_container_layout = QVBoxLayout(self.task_container)
        self.task_container_layout.setContentsMargins(24, 4, 0, 0)
        self.task_container_layout.setSpacing(2)

        for task in self.tasks:
            item = TaskItem(task, self.on_toggle, self._check_changed)
            self.task_container_layout.addWidget(item)

        main_layout.addWidget(header)
        main_layout.addWidget(self.task_container)

        # 非今日默认折叠
        if self.collapsed:
            self.task_container.setVisible(False)
            self.toggle_btn.setText("▶")

    def _toggle(self):
        self.collapsed = not self.collapsed
        self.task_container.setVisible(not self.collapsed)
        self.toggle_btn.setText("▼" if not self.collapsed else "▶")


# ==================== Task Item Widget ====================

class TaskItem(QFrame):
    def __init__(self, task_data, on_toggle, on_check_changed=None, parent=None):
        super().__init__(parent)
        self.task_data = task_data
        self.on_toggle = on_toggle
        self.on_check_changed = on_check_changed  # 回调用于通知父区块更新统计
        self.setup_ui()

    def setup_ui(self):
        self.setStyleSheet("background: transparent;")
        self.setFixedHeight(52)
        self.setMinimumWidth(300)
        self.setObjectName("taskItem")

        # 初始化滚动相关属性（需要在_apply_done_style之前）
        self._marquee_text = self.task_data["task"]
        self._marquee_offset = 0
        self._marquee_timer = None

        layout = QHBoxLayout(self)
        layout.setContentsMargins(14, 6, 14, 6)
        layout.setSpacing(12)

        # Checkbox
        self.cb = QCheckBox()
        self.cb.setStyleSheet("""
            QCheckBox { spacing: 10px; }
            QCheckBox::indicator { width: 20px; height: 20px; border-radius: 6px; border: 1px solid rgba(100,80,180,100); background: rgba(30,30,50,120); }
            QCheckBox::indicator:checked { background: rgba(80,180,120,60); border-color: rgba(80,200,140,120); }
        """)
        self.cb.stateChanged.connect(self._on_checked)

        # Task info（需要先创建label，因为_apply_done_style依赖它）
        info_layout = QVBoxLayout()
        info_layout.setSpacing(1)

        self.task_label = QLabel(self.task_data["task"])
        self.task_label.setStyleSheet("color: white; font-size: 13px;")
        self.task_label.setWordWrap(False)
        self.task_label.setFixedHeight(18)
        self.task_label.setTextInteractionFlags(Qt.NoTextInteraction)  # 禁用选中

        # 检查文字宽度，决定是否需要滚动
        fm = self.task_label.fontMetrics()
        text_width = fm.horizontalAdvance(self.task_data["task"])
        avail_width = 260  # 减去checkbox和边距后的估算宽度
        self._needs_marquee = text_width > avail_width

        # 鼠标悬停/离开事件
        self.task_label.setMouseTracking(True)
        self.task_label.enterEvent = self._on_mouse_enter
        self.task_label.leaveEvent = self._on_mouse_leave

        self.type_label = QLabel(self.task_data["type"])
        type_color = get_type_color(self.task_data["type"])
        self.type_label.setStyleSheet(f"color: {type_color}; font-size: 10px; background: rgba(255,255,255,0.06); border-radius: 4px; padding: 1px 6px;")

        info_layout.addWidget(self.task_label)
        info_layout.addWidget(self.type_label)

        # 初始化已完成任务的状态（在所有UI创建完成后）
        if self.task_data.get("done"):
            self.cb.blockSignals(True)
            self.cb.setChecked(True)
            self.cb.blockSignals(False)
            self._apply_done_style()

        layout.addWidget(self.cb, 0, Qt.AlignTop)
        layout.addLayout(info_layout, 1)

    def _start_marquee(self):
        if self._marquee_timer:
            return
        from PySide6.QtCore import QTimer
        self._marquee_timer = QTimer(self)
        self._marquee_timer.timeout.connect(self._scroll_text)
        self._marquee_timer.setInterval(120)  # 慢速滚动
        self._marquee_timer.start()
        self.task_label.setText(self._marquee_text)

    def _scroll_text(self):
        if not hasattr(self, '_marquee_offset'):
            return
        self._marquee_offset += 1
        text = self._marquee_text
        full_text = text + "    " + text
        step = self._marquee_offset % (len(text) + 4)
        self.task_label.setText(full_text[step:step+20])
        self.task_label.setToolTip(text)

    def _stop_marquee(self):
        if hasattr(self, '_marquee_timer') and self._marquee_timer:
            self._marquee_timer.stop()
            self._marquee_timer.deleteLater()
            self._marquee_timer = None

    def _on_mouse_enter(self, event):
        if self._needs_marquee and not self.cb.isChecked():
            self._start_marquee()

    def _on_mouse_leave(self, event):
        self._stop_marquee()
        self.task_label.setText(self._marquee_text)

    def _apply_done_style(self):
        """应用已完成样式（绿色删除线）"""
        self._stop_marquee()
        self.task_label.setText(self._marquee_text)
        self.task_label.setStyleSheet("color: rgba(16,185,129,0.9); font-size: 13px; text-decoration: line-through;")
        self.cb.setStyleSheet("""
            QCheckBox { spacing: 8px; }
            QCheckBox::indicator { width: 18px; height: 18px; border-radius: 4px; border: 1px solid rgba(16,185,129,0.6); background: rgba(16,185,129,0.2); }
            QCheckBox::indicator:checked { background: rgba(16,185,129,0.3); border-color: rgba(16,185,129,0.6); }
        """)

    def _set_saving_state(self):
        """正在改写Excel时的中间状态"""
        self._stop_marquee()
        self.task_label.setText(self._marquee_text)
        self.task_label.setStyleSheet("color: rgba(251,191,36,0.85); font-size: 13px;")
        self.cb.setStyleSheet("""
            QCheckBox { spacing: 8px; }
            QCheckBox::indicator { width: 18px; height: 18px; border-radius: 4px; border: 1px solid rgba(251,191,36,0.6); background: rgba(251,191,36,0.15); animation: pulse 1s infinite; }
            QCheckBox::indicator:checked { background: rgba(124,58,237,0.4); border-color: #7C3AED; }
        """)

    def _on_checked(self, state):
        done = (state == 2)  # PySide6: stateChanged发射的state是整数2代表Checked
        if done:
            self._set_saving_state()
        else:
            # 取消完成时恢复到白色
            self.task_label.setStyleSheet("color: white; font-size: 13px; text-decoration: none;")
            self.cb.setStyleSheet("""
                QCheckBox { spacing: 8px; }
                QCheckBox::indicator { width: 18px; height: 18px; border-radius: 4px; border: 1px solid rgba(255,255,255,0.2); background: transparent; }
                QCheckBox::indicator:checked { background: rgba(124,58,237,0.4); border-color: #7C3AED; }
            """)
        self.on_toggle(self.task_data["row_num"], done, self)
        if self.on_check_changed:
            self.on_check_changed(done)

    def mark_synced(self, state):
        """
        state: "done" | "undo" | "saving"
        - done: 已完成（绿色+删除线）
        - undo: 未完成（白色正常）
        - saving: 正在保存（黄色）
        """
        self.cb.blockSignals(True)
        self.cb.setChecked(state != "undo")
        self.cb.blockSignals(False)
        self._stop_marquee()
        self.task_label.setText(self._marquee_text)

        if state == "done":
            self.task_label.setStyleSheet("color: rgba(16,185,129,0.9); font-size: 13px; text-decoration: line-through;")
            self.setFixedHeight(46)
            self.cb.setStyleSheet("""
                QCheckBox { spacing: 8px; }
                QCheckBox::indicator { width: 18px; height: 18px; border-radius: 4px; border: 1px solid rgba(16,185,129,0.6); background: rgba(16,185,129,0.2); }
                QCheckBox::indicator:checked { background: rgba(16,185,129,0.3); border-color: rgba(16,185,129,0.6); }
            """)
        elif state == "saving":
            self.task_label.setStyleSheet("color: rgba(251,191,36,0.9); font-size: 13px;")
            self.cb.setStyleSheet("""
                QCheckBox { spacing: 8px; }
                QCheckBox::indicator { width: 18px; height: 18px; border-radius: 4px; border: 1px solid rgba(251,191,36,0.6); background: rgba(251,191,36,0.15); }
                QCheckBox::indicator:checked { background: rgba(251,191,36,0.2); border-color: rgba(251,191,36,0.6); }
            """)
        else:  # undo
            self.task_label.setStyleSheet("color: white; font-size: 13px; text-decoration: none;")
            self.setFixedHeight(48)
            self.cb.setStyleSheet("""
                QCheckBox { spacing: 8px; }
                QCheckBox::indicator { width: 18px; height: 18px; border-radius: 4px; border: 1px solid rgba(255,255,255,0.2); background: transparent; }
                QCheckBox::indicator:checked { background: rgba(124,58,237,0.4); border-color: #7C3AED; }
            """)

    def deleteLater(self):
        self._stop_marquee()
        super().deleteLater()

# ==================== Main Application ====================

class DailyPlanApp(QWidget):
    WINDOW_WIDTH = 360  # 正常模式宽度
    WINDOW_HEIGHT = 520
    COMPACT_HEIGHT = 24  # 紧凑模式高度
    COMPACT_WIDTH = 280  # 紧凑模式宽度
    BG_COLOR = QColor(35, 35, 65, 255)  # 渐变起始色（深紫）
    BG_COLOR_END = QColor(15, 55, 95, 255)  # 渐变结束色（深蓝）
    BG_ALERT = QColor(180, 20, 30, 220)

    # 信号：ok表示是否保存成功，done_ref表示原本意图是完成还是取消，widget是要更新的组件
    _thread_callback = Signal(object, object, object)

    def __init__(self):
        super().__init__()
        self.compact_mode = False
        self.normal_geometry = QRect(0, 0, self.WINDOW_WIDTH, self.WINDOW_HEIGHT)
        self.drag_pos = None
        self._blur_set = False
        self._suppress_compact = False
        self._press_pos = QPoint()
        self._pulse_opacity = 0.0
        self.tasks = []

        # 连接信号到槽
        self._thread_callback.connect(self._on_thread_result)

        self.setup_window()
        self.setup_ui()
        self.setup_refresh_timer()
        self.reload_tasks()

    # ────────── Window Setup ──────────

    def setup_window(self):
        self.setWindowFlags(
            Qt.FramelessWindowHint |
            Qt.WindowStaysOnTopHint |
            Qt.Window
        )
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedSize(self.WINDOW_WIDTH, self.WINDOW_HEIGHT)
        self.setMouseTracking(True)

    def showEvent(self, event):
        super().showEvent(event)
        if not self._blur_set and self.winId():
            # 移除 acrylic blur（它不支持真正的透明背景）
            # enable_acrylic_blur(int(self.winId()))
            self._blur_set = True

    # ────────── UI Setup ──────────

    def setup_ui(self):
        # ── Top bar ──
        self.top_bar = QHBoxLayout()
        self.top_bar.setContentsMargins(14, 8, 14, 0)

        self.icon_label = QLabel("📋")
        self.icon_label.setStyleSheet("font-size: 15px;")

        self.title_label = QLabel("每日学习计划")
        self.title_label.setStyleSheet("font-size: 12px; color: rgba(255,255,255,0.5); letter-spacing: 1px;")

        self.top_bar.addWidget(self.icon_label)
        self.top_bar.addWidget(self.title_label)
        self.top_bar.addStretch()

        # Refresh button
        self.refresh_btn = ModernButton("↻", padding="4px 10px", font_size="14px", color="rgba(255,255,255,0.5)")
        self.refresh_btn.setFixedSize(26, 24)
        self.refresh_btn.clicked.connect(self.reload_tasks)
        self.refresh_btn.setToolTip("刷新任务")

        # Close button
        self.close_btn = ModernButton("✕", bg_hover="rgba(220,40,40,0.6)", padding="4px 10px", font_size="14px", color="rgba(255,255,255,0.5)")
        self.close_btn.setFixedSize(26, 24)
        self.close_btn.clicked.connect(self.close)
        self.close_btn.setToolTip("关闭")

        self.top_bar.addWidget(self.refresh_btn)
        self.top_bar.addWidget(self.close_btn)

        # ── Compact mode label (hidden by default) ──
        self.compact_label = QLabel()
        self.compact_label.setFixedHeight(self.COMPACT_HEIGHT)
        self.compact_label.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
        self.compact_label.setFont(QFont("Consolas", 14, QFont.Weight.Normal))
        self.compact_label.setStyleSheet("color: rgba(255,255,255,0.7); font-size: 14px; background: transparent;")
        self.compact_label.hide()

        # ── Task scroll area ──
        self.scroll = QScrollArea()
        self.scroll.setStyleSheet("""
            QScrollArea { background: transparent; border: none; }
            QScrollBar:vertical { background: rgba(255,255,255,0.05); width: 4px; border-radius: 2px; }
            QScrollBar::handle { background: rgba(255,255,255,0.1); border-radius: 2px; }
        """)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll.setWidgetResizable(True)

        self.task_container = QFrame()
        self.task_container.setStyleSheet("background: transparent;")
        self.task_layout = QVBoxLayout(self.task_container)
        self.task_layout.setContentsMargins(0, 0, 0, 0)
        self.task_layout.setSpacing(4)
        self.task_layout.addStretch()

        self.scroll.setWidget(self.task_container)

        # ── Bottom bar ──
        self.bottom_bar = QHBoxLayout()
        self.bottom_bar.setContentsMargins(14, 8, 14, 0)

        self.stats_label = QLabel("加载中...")
        self.stats_label.setStyleSheet("color: rgba(255,255,255,0.4); font-size: 11px;")

        self.open_excel_btn = ModernButton("打开 Excel", padding="6px 12px", font_size="11px")
        self.open_excel_btn.clicked.connect(self._open_excel)
        self.open_excel_btn.setStyleSheet(self.open_excel_btn.styleSheet().replace("rgba(255,255,255,0.15)", "rgba(124,58,237,0.3)").replace("rgba(255,255,255,0.3)", "rgba(124,58,237,0.5)"))

        self.top_bar_widget = QFrame()
        self.top_bar_widget.setLayout(self.top_bar)

        self.bottom_bar_widget = QFrame()
        self.bottom_bar_widget.setLayout(self.bottom_bar)

        self.bottom_bar.addWidget(self.stats_label)
        self.bottom_bar.addStretch()
        self.bottom_bar.addWidget(self.open_excel_btn)

        # ── Main layout ──
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(20, 0, 20, 20)
        self.main_layout.setSpacing(2)

        self.main_layout.addWidget(self.top_bar_widget)
        self.main_layout.addSpacing(6)
        self.main_layout.addWidget(self.compact_label)
        self.main_layout.addWidget(self.scroll, 1)
        self.main_layout.addSpacing(6)
        self.main_layout.addWidget(self.bottom_bar_widget)

    def setup_refresh_timer(self):
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self.reload_tasks)
        self.refresh_timer.setInterval(REFRESH_INTERVAL_MS)

    # ────────── Task Loading ──────────

    def reload_tasks(self):
        self.tasks = load_tasks()
        self.refresh_task_list()

    def refresh_task_list(self):
        # Remove existing task widgets
        while self.task_layout.count() > 1:
            item = self.task_layout.takeAt(0)
            if item.widget() and item.widget() != self.task_layout.itemAt(self.task_layout.count() - 1).widget():
                item.widget().deleteLater()

        if not self.tasks:
            # Show empty state
            empty = QLabel("暂无今日任务")
            empty.setAlignment(Qt.AlignCenter)
            empty.setStyleSheet("color: rgba(255,255,255,0.3); font-size: 13px; padding: 20px;")
            self.task_layout.insertWidget(0, empty)
            self.stats_label.setText("暂无任务")
            return

        # Group tasks by date
        from collections import defaultdict
        date_groups = defaultdict(list)
        for task in self.tasks:
            date_groups[task["date"]].append(task)

        # Sort by date (earliest first)
        for date_key in sorted(date_groups.keys()):
            tasks_for_date = date_groups[date_key]
            month_day = date_key.strftime("%m月%d日")
            section = DateSection(month_day, tasks_for_date, self._on_task_toggle, is_today=(date_key == date.today()))
            # Keep reference so stats can be updated
            self.task_layout.insertWidget(self.task_layout.count() - 1, section)

        total = len(self.tasks)
        done = sum(1 for t in self.tasks if t["done"])
        self.stats_label.setText(f"共 {total} 项 · 已完成 {done}")

    def _on_task_toggle(self, row_num, done, widget):
        # 先更新为saving状态
        widget.mark_synced("saving")
        for t in self.tasks:
            if t["row_num"] == row_num:
                t["done"] = done
                break
        total = len(self.tasks)
        done_count = sum(1 for t in self.tasks if t["done"])
        self.stats_label.setText(f"共 {total} 项 · 已完成 {done_count}")

        # 保存widget引用和done状态，用于线程回调
        widget_ref = widget
        done_ref = done

        # 后台线程写Excel，写完后通过信号回调主线程
        def do_save():
            ok = mark_done(row_num, done)
            # 用信号把结果传回主线程（信号机制保证跨线程安全）
            self._thread_callback.emit(ok, done_ref, widget_ref)

        threading.Thread(target=do_save, daemon=True).start()

    def _on_thread_result(self, ok, done_ref, widget):
        """信号槽回调，在主线程执行UI更新"""
        if ok:
            widget.mark_synced("done" if done_ref else "undo")
        else:
            widget.mark_synced("undo")

    # ────────── Excel Integration ──────────

    def _open_excel(self):
        import subprocess
        subprocess.Popen(["start", "", EXCEL_PATH], shell=True)

    # ────────── Compact Mode ──────────

    def enter_compact_mode(self):
        if self.compact_mode:
            return
        self.compact_mode = True
        self.normal_geometry = self.geometry()

        self._normal_margins = self.main_layout.contentsMargins()
        # 紧凑模式：上下左右都收紧
        self.main_layout.setContentsMargins(2, 0, 2, 0)
        self.top_bar.setContentsMargins(8, 4, 8, 0)
        self.bottom_bar.setContentsMargins(8, 4, 8, 0)

        x = self.normal_geometry.x()
        self.setFixedSize(self.COMPACT_WIDTH, self.COMPACT_HEIGHT)
        self.move(x, 0)

        self.scroll.hide()
        self.bottom_bar_widget.hide()
        self.top_bar_widget.hide()
        self.compact_label.show()

        self.compact_label.setFont(QFont("Consolas", 14, QFont.Weight.Normal))
        self.compact_label.setFixedHeight(self.COMPACT_HEIGHT)
        self.compact_label.setAlignment(Qt.AlignTop | Qt.AlignHCenter)
        self.compact_label.setGraphicsEffect(None)

        self._update_compact_display()

    def exit_compact_mode(self):
        if not self.compact_mode:
            return
        self.compact_mode = False

        if hasattr(self, '_normal_margins'):
            self.main_layout.setContentsMargins(self._normal_margins)
            self.top_bar.setContentsMargins(14, 8, 14, 0)
            self.bottom_bar.setContentsMargins(14, 8, 14, 0)

        self.setFixedHeight(self.WINDOW_HEIGHT)
        self.setFixedWidth(self.WINDOW_WIDTH)
        if self.normal_geometry:
            geo = self.normal_geometry
            # 确保 y 不为负数，顶部与屏幕顶部对齐
            if geo.top() < 0:
                geo = QRect(geo.left(), 0, geo.width(), geo.height())
            self.setGeometry(geo)

        self.scroll.show()
        self.bottom_bar_widget.show()
        self.top_bar_widget.show()
        self.compact_label.hide()

        self.compact_label.setFont(QFont("Consolas", 14, QFont.Weight.Normal))
        self.compact_label.setFixedHeight(self.COMPACT_HEIGHT)
        self.compact_label.setAlignment(Qt.AlignTop | Qt.AlignHCenter)

    def _set_compact_widgets_visible(self, visible):
        pass  # managed via explicit show/hide

    def _update_compact_display(self):
        # 今日任务：今天日期的任务完成情况
        today = date.today()
        today_tasks = [t for t in self.tasks if t["date"] == today]
        today_done = sum(1 for t in today_tasks if t["done"])
        today_total = len(today_tasks)

        # 待完成任务：所有未完成的任务
        pending = sum(1 for t in self.tasks if not t["done"])

        self.compact_label.setText(f"今日任务({today_done}/{today_total}) 待完成任务:{pending}")

    # ────────── Mouse Events ──────────

    def mousePressEvent(self, event: QMouseEvent):
        if event.button() == Qt.LeftButton:
            self.drag_pos = event.globalPosition().toPoint()
            if self.compact_mode:
                self._press_pos = event.globalPosition().toPoint()

    def mouseMoveEvent(self, event: QMouseEvent):
        if event.buttons() & Qt.LeftButton and self.drag_pos is not None:
            if self.compact_mode:
                delta = event.globalPosition().toPoint() - self._press_pos
                if abs(delta.y()) > 15:
                    old_x = self.normal_geometry.x() if self.normal_geometry else self.x()
                    self.exit_compact_mode()
                    cursor_y = int(event.globalPosition().y())
                    self.move(old_x, max(0, cursor_y - 40))
                    self.normal_geometry = QRect(self.x(), self.y(), self.WINDOW_WIDTH, self.WINDOW_HEIGHT)
                    self.drag_pos = event.globalPosition().toPoint()
                    self._suppress_compact = True
                    QTimer.singleShot(400, lambda: setattr(self, '_suppress_compact', False))
                return

            delta = event.globalPosition().toPoint() - self.drag_pos
            self.move(self.pos() + delta)
            self.drag_pos = event.globalPosition().toPoint()

    def mouseReleaseEvent(self, event: QMouseEvent):
        if event.button() == Qt.LeftButton:
            if self.compact_mode:
                self.exit_compact_mode()
                self.drag_pos = None
                return
            self.drag_pos = None
            if not self._suppress_compact and self.pos().y() <= 5:
                self.enter_compact_mode()

    def mouseDoubleClickEvent(self, event: QMouseEvent):
        if event.button() == Qt.LeftButton:
            if self.compact_mode:
                self.exit_compact_mode()

    # ────────── Painting ──────────

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # 获取绘制区域
        r = self.rect()

        if self.compact_mode:
            # 紧凑模式：纯色填充，四角圆角
            corner = 12
        else:
            corner = 16

        path = QPainterPath()
        path.addRoundedRect(QRectF(r), corner, corner)
        painter.fillPath(path, self.BG_COLOR)
        painter.setPen(QPen(QColor(100, 80, 180, 80), 1.5))
        painter.drawPath(path)


# ==================== Main Entry ====================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setApplicationName("每日学习计划")

    palette = app.palette()
    palette.setColor(palette.ColorRole.WindowText, Qt.white)
    palette.setColor(palette.ColorRole.Text, Qt.white)
    app.setPalette(palette)

    window = DailyPlanApp()
    window.setWindowIcon(QIcon("D:/Code_Project/Claude_code_project/demo/app.ico"))
    screen = app.primaryScreen().availableGeometry()
    window.move(
        (screen.width() - window.width()) // 2,
        (screen.height() - window.height()) // 2
    )
    window.show()

    sys.exit(app.exec())